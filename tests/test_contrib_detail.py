"""Verification detaillee Contrib + comparaison export vs original."""
import io, os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import openpyxl, pandas as pd, numpy as np
from config import CONSOLIDES
from io_utils.excel_reader import read_consignes, read_codification, read_donnees_calcul, rename_columns_to_codes
from io_utils.excel_writer import write_icae_recalc_output
from core.icae_engine import run_icae_pipeline
from core.quarterly import quarterly_mean, calc_ga_trim, calc_gt_trim, contributions_sectorielles_trim, normalize_contrib_to_ga

WS = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
fp = os.path.join(WS, "Livrable_Final", "01_Classeurs_ICAE", "ICAE_RCA_Consolide_now.xlsx")
codif = read_codification(fp)
donnees = read_donnees_calcul(fp)
donnees = rename_columns_to_codes(donnees, codif)
consignes = read_consignes(fp)
base_year = consignes.get("base_year", 2023)

last_date = pd.to_datetime(donnees["Date"]).max()
fcst_dates = pd.date_range(last_date + pd.DateOffset(months=1), periods=3, freq="MS")
new_rows = pd.DataFrame({"Date": fcst_dates})
for c in [x for x in donnees.columns if x != "Date"]:
    lv = donnees[c].dropna().iloc[-1] if donnees[c].notna().any() else 100
    new_rows[c] = [lv * (1 + np.random.randn() * 0.01) for _ in range(3)]
extended = pd.concat([donnees, new_rows], ignore_index=True)

dates_c = pd.to_datetime(extended["Date"])
base_idx = extended.index[dates_c.dt.year == base_year]
base_rows = range(base_idx[0], base_idx[-1] + 1)
priors = pd.Series(codif["PRIOR"].values, index=codif["Code"].values, dtype=float).fillna(0)
results = run_icae_pipeline(donnees=extended, priors=priors, base_year=base_year, base_rows=base_rows)
q = quarterly_mean(results["icae"], results["dates"])
q["GA_Trim"] = calc_ga_trim(q["icae_trim"])
q["GT_Trim"] = calc_gt_trim(q["icae_trim"])
contrib_trim = contributions_sectorielles_trim(results["weights"]["m"], codif, results["dates"])
contrib_trim = normalize_contrib_to_ga(contrib_trim, q["GA_Trim"])

wb_bytes = write_icae_recalc_output(donnees=extended, results=results, quarterly=q,
                                     contrib_trim=contrib_trim, codification=codif,
                                     country_code="RCA", source_path=fp)
wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
wb_orig = openpyxl.load_workbook(fp)

print("=== VERIFICATION DETAILLEE ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    orig_ws = wb_orig[sn] if sn in wb_orig.sheetnames else None
    orig_info = f"orig={orig_ws.max_row}" if orig_ws else "N/A"
    print(f"  {sn}: export={ws.max_row} rows, {orig_info}")

ws_ct = wb["Contrib"]
print(f"\nContrib headers: {[ws_ct.cell(1,j).value for j in range(1, ws_ct.max_column+1)]}")
print(f"contrib_trim shape: {contrib_trim.shape}")
print(f"Contrib export rows: {ws_ct.max_row}")
ws_co = wb_orig["Contrib"]
print(f"Contrib orig  rows: {ws_co.max_row}")
print(f"Attendu: {len(contrib_trim) + 1}")

print("\nContrib last 3 rows export:")
for r in range(max(2, ws_ct.max_row - 2), ws_ct.max_row + 1):
    vals = [ws_ct.cell(r, j).value for j in range(1, min(ws_ct.max_column + 1, 6))]
    print(f"  R{r}: {vals}")

ws_d = wb["Donnees_calcul"]
ws_do = wb_orig["Donnees_calcul"]
print(f"\nDonnees_calcul export: {ws_d.max_row} | orig: {ws_do.max_row} | attendu: {len(extended)+1}")

ws_c = wb["CALCUL_ICAE"]
ws_co2 = wb_orig["CALCUL_ICAE"]
print(f"CALCUL_ICAE   export: {ws_c.max_row} | orig: {ws_co2.max_row}")

ws_rt = wb["Resultats_Trim"]
ws_rto = wb_orig["Resultats_Trim"]
print(f"Resultats_Trim export: {ws_rt.max_row} | orig: {ws_rto.max_row} | attendu: {len(q)+1}")

# Verifier que les NOUVELLES lignes de Donnees_calcul ont des formattages
print("\nDonnees_calcul - 5 dernieres lignes export:")
for r in range(ws_d.max_row - 4, ws_d.max_row + 1):
    d = ws_d.cell(r, 1).value
    v = ws_d.cell(r, 2).value
    fill = ws_d.cell(r, 2).fill
    c = fill.start_color.rgb if fill and fill.patternType else "no-fill"
    print(f"  R{r}: date={d}, v2={v}, fill={c}")

print("\nCALCUL_ICAE - 5 dernieres lignes export:")
for r in range(ws_c.max_row - 4, ws_c.max_row + 1):
    a = ws_c.cell(r, 1).value
    b = ws_c.cell(r, 2).value
    print(f"  R{r}: A={str(a)[:40]}, B={str(b)[:50]}")

wb.close()
wb_orig.close()
print("\n=== OK ===")
