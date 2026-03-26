"""Read weights from reference CEMAC file."""
import openpyxl
import sys

path = r"c:\Users\HP\Documents\Stage pro BEAC\Work\ICAE\Mars 2026\Livrable_Final\01_Classeurs_ICAE\ICAE_CEMAC_Consolide.xlsx"
try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    ws = wb["Poids_PIB"]
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print(row)
    wb.close()
except Exception as e:
    print(f"Error: {e}", file=sys.stderr)
