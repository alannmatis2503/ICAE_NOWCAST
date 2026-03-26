"""Verification rapide des colonnes sectorielles dans Contrib export."""
import io, os, sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
import openpyxl, pandas as pd, numpy as np
from io_utils.excel_reader import read_consignes, read_codification, read_donnees_calcul, rename_columns_to_codes
from io_utils.excel_writer import write_icae_recalc_output
from core.icae_engine import run_icae_pipeline
from core.quarterly import quarterly_mean, calc_ga_trim, calc_gt_trim, contributions_sectorielles_trim, normalize_contrib_to_ga

WS = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
fp = os.path.join(WS, "Livrable_Final", "01_Classeurs_ICAE", "ICAE_RCA_Consolide_now.xlsx")
codif = read_codification(fp)
donnees = read_donnees_calcul(fp)
donnees = rename_columns_to_codes(donnees, codif)
consignes = read_consignes(fp)
base_year = consignes.get("base_year", 2023)

last_date = pd.to_datetime(donnees["Date"]).max()
fcst_dates = pd.date_range(last_date + pd.DateOffset(months=1), periods=3, freq="MS")
new_rows = pd.DataFrame({"Date": fcst_dates})
for c in [x for x in donnees.columns if x != "Date"]:
    lv = donnees[c].dropna().iloc[-1] if donnees[c].notna().any() else 100
    new_rows[c] = [lv * (1 + np.random.randn() * 0.01) for _ in range(3)]
extended = pd.concat([donnees, new_rows], ignore_index=True)

dates_c = pd.to_datetime(extended["Date"])
base_idx = extended.index[dates_c.dt.year == base_year]
base_rows = range(base_idx[0], base_idx[-1] + 1)
priors = pd.Series(codif["PRIOR"].values, index=codif["Code"].values, dtype=float).fillna(0)
results = run_icae_pipeline(donnees=extended, priors=priors, base_year=base_year, base_rows=base_rows)
q = quarterly_mean(results["icae"], results["dates"])
q["GA_Trim"] = calc_ga_trim(q["icae_trim"])
q["GT_Trim"] = calc_gt_trim(q["icae_trim"])
contrib_trim = contributions_sectorielles_trim(results["weights"]["m"], codif, results["dates"])
contrib_trim = normalize_contrib_to_ga(contrib_trim, q["GA_Trim"])

wb_bytes = write_icae_recalc_output(donnees=extended, results=results, quarterly=q,
                                     contrib_trim=contrib_trim, codification=codif,
                                     country_code="RCA", source_path=fp)
wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
ws_ct = wb["Contrib"]

print("=== CONTRIB: toutes les colonnes des 3 dernieres lignes ===")
for r in range(ws_ct.max_row - 2, ws_ct.max_row + 1):
    print(f"\nRow {r}:")
    for c in range(1, ws_ct.max_column + 1):
        v = ws_ct.cell(r, c).value
        hdr = ws_ct.cell(1, c).value
        print(f"  Col {c} ({str(hdr)[:20] if hdr else '-'}): {v}")

print("\n=== Comparison row 2 vs new rows (pattern check) ===")
for c in range(24, ws_ct.max_column + 1):
    v2 = ws_ct.cell(2, c).value
    v_new = ws_ct.cell(ws_ct.max_row, c).value
    hdr = ws_ct.cell(1, c).value
    print(f"Col {c} ({str(hdr)[:20]}): R2={v2}  | New={v_new}")

wb.close()
print("\n=== OK ===")
