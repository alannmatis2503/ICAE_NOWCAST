"""
Test end-to-end : CEMAC + Module 2 recalc
Reproduit exactement le flux de l'application Streamlit.
"""
import sys, os, io, traceback

# ── Setup chemin ─────────────────────────────────────────────────────────
APP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, APP_DIR)
WS_ROOT = os.path.abspath(os.path.join(APP_DIR, "..", ".."))

import pandas as pd
import numpy as np
import openpyxl

from config import (COUNTRY_CODES, COUNTRY_NAMES, PIB_2014, POIDS_PIB,
                    CEMAC_TEMPLATE, CONSOLIDES)
from core.cemac_engine import compute_icae_cemac, quarterly_cemac
from core.icae_engine import run_icae_pipeline
from io_utils.excel_reader import (
    list_sheets, read_consignes, read_codification,
    read_donnees_calcul, rename_columns_to_codes,
)
from io_utils.excel_writer import write_cemac_excel, write_icae_recalc_output
from core.quarterly import (
    quarterly_mean, calc_ga_trim, calc_gt_trim,
    contributions_sectorielles_trim, normalize_contrib_to_ga,
)

errors = []

# =====================================================================
# PARTIE 1 : Test CEMAC end-to-end (fichiers Livrable_Final)
# =====================================================================
print("=" * 70)
print("PARTIE 1 : Test CEMAC end-to-end")
print("=" * 70)

LIVRABLE_DIR = os.path.join(WS_ROOT, "Livrable_Final", "01_Classeurs_ICAE")
print(f"Dossier source : {LIVRABLE_DIR}")
print(f"Existe : {os.path.isdir(LIVRABLE_DIR)}")

# 1a) Charger chaque fichier pays (reproduit _process_country_file)
icae_dict = {}
dates_dict = {}

for code in COUNTRY_CODES:
    # Trouver le fichier
    patterns = [
        f"ICAE_{code}_Consolide_now.xlsx",
        f"ICAE_{code}_Consolide.xlsx",
    ]
    fpath = None
    for pat in patterns:
        candidate = os.path.join(LIVRABLE_DIR, pat)
        if os.path.exists(candidate):
            fpath = candidate
            break

    if fpath is None:
        print(f"  [{code}] MANQUANT - aucun fichier trouvé")
        continue

    print(f"\n  [{code}] Fichier : {os.path.basename(fpath)}")
    try:
        sheets = list_sheets(fpath)
        print(f"    Feuilles : {sheets}")

        if "Donnees_calcul" not in sheets:
            print(f"    SKIP : pas de Donnees_calcul")
            continue

        consignes = read_consignes(fpath) if "Consignes" in sheets else {"base_year": 2023}
        codif = read_codification(fpath) if "Codification" in sheets else pd.DataFrame()
        donnees = read_donnees_calcul(fpath)
        print(f"    Donnees_calcul : {len(donnees)} lignes, {len(donnees.columns)} colonnes")

        if not codif.empty:
            donnees = rename_columns_to_codes(donnees, codif)
            print(f"    Colonnes renommées : {list(donnees.columns[:5])}...")

        base_year = consignes.get("base_year", 2023)
        dates = pd.to_datetime(donnees["Date"])
        base_mask = dates.dt.year == base_year
        base_indices = donnees.index[base_mask]
        if len(base_indices) > 0:
            base_rows = range(base_indices[0], base_indices[-1] + 1)
        else:
            bs = consignes.get("base_rows_start", 124) - 16
            be = consignes.get("base_rows_end", 135) - 16
            base_rows = range(bs, be + 1)

        if not codif.empty and "Code" in codif.columns and "PRIOR" in codif.columns:
            priors = pd.Series(
                codif["PRIOR"].values, index=codif["Code"].values, dtype=float,
            ).fillna(0)
        else:
            data_cols = [c for c in donnees.columns if c != "Date"]
            priors = pd.Series(1.0, index=data_cols)

        results = run_icae_pipeline(
            donnees=donnees, priors=priors,
            base_year=base_year, base_rows=base_rows,
        )
        icae = results["icae"]
        icae.index = dates
        icae_dict[code] = icae
        dates_dict[code] = dates
        print(f"    ICAE : {len(icae)} obs, last={icae.dropna().iloc[-1]:.2f}")
    except Exception as e:
        print(f"    ERREUR : {e}")
        traceback.print_exc()
        errors.append(f"CEMAC load {code}: {e}")

print(f"\n  Pays chargés : {list(icae_dict.keys())}")

# 1b) Filtrage période commune
if icae_dict:
    country_ranges = {}
    for code, series in icae_dict.items():
        idx = pd.to_datetime(series.index)
        valid = idx[series.notna()]
        if len(valid) > 0:
            country_ranges[code] = (valid.min(), valid.max())

    common_start = max(r[0] for r in country_ranges.values())
    common_end = min(r[1] for r in country_ranges.values())
    print(f"\n  Période commune : {common_start} - {common_end}")

    # Filtrer
    filtered_dict = {}
    for code, series in icae_dict.items():
        idx = pd.to_datetime(series.index)
        mask = (idx >= common_start) & (idx <= common_end)
        filtered_dict[code] = series[mask]
        print(f"    {code}: {len(series[mask])} obs après filtrage")

    # 1c) compute_icae_cemac
    print("\n  Calcul ICAE CEMAC...")
    try:
        result_df = compute_icae_cemac(filtered_dict, POIDS_PIB)
        print(f"    result_df : {result_df.shape}")
        print(f"    Colonnes : {list(result_df.columns)}")
        print(f"    Index type : {type(result_df.index)}")
        print(f"    ICAE CEMAC last : {result_df['ICAE_CEMAC'].dropna().iloc[-1]:.2f}")
    except Exception as e:
        print(f"    ERREUR compute_icae_cemac : {e}")
        traceback.print_exc()
        errors.append(f"compute_icae_cemac: {e}")
        result_df = pd.DataFrame()

    # 1d) Filtrage graphique (reproduit le code qui posait problème)
    if not result_df.empty:
        print("\n  Test filtrage graphique (lignes 283-286)...")
        try:
            _res_dates = pd.to_datetime(result_df.index)
            _graph_start = pd.Timestamp(_res_dates.min())
            _graph_end = pd.Timestamp(_res_dates.max())

            _graph_mask = (_res_dates >= _graph_start) & (_res_dates <= _graph_end)
            print(f"    Type de _graph_mask : {type(_graph_mask)}")
            print(f"    _graph_mask dtype : {_graph_mask.dtype}")

            # C'est ICI que le bug se manifestait :
            # result_df[_graph_mask.values]  <-- AttributeError
            _disp_df = result_df[_graph_mask]  # fix : pas de .values
            print(f"    _disp_df : {_disp_df.shape} lignes -- OK !")
        except Exception as e:
            print(f"    ERREUR filtrage : {e}")
            traceback.print_exc()
            errors.append(f"graph_mask: {e}")

    # 1e) Quarterly
    if not result_df.empty:
        print("\n  Calcul trimestriel...")
        try:
            dates_idx = pd.to_datetime(result_df.index)
            q_cemac = quarterly_cemac(result_df, dates_idx)
            print(f"    q_cemac : {q_cemac.shape}")
            if "Trimestre" in q_cemac.columns:
                q_cemac["Trimestre"] = q_cemac["Trimestre"].astype(str)
            print(f"    Derniers trimestres : {list(q_cemac['Trimestre'].tail(4))}")
        except Exception as e:
            print(f"    ERREUR quarterly_cemac : {e}")
            traceback.print_exc()
            errors.append(f"quarterly_cemac: {e}")
            q_cemac = pd.DataFrame()

    # 1f) Export CEMAC (from scratch)
    if not result_df.empty and not q_cemac.empty:
        print("\n  Export CEMAC (from scratch)...")
        try:
            data_bytes = write_cemac_excel(result_df, q_cemac, POIDS_PIB)
            print(f"    OK : {len(data_bytes)} bytes")
        except Exception as e:
            print(f"    ERREUR export scratch : {e}")
            traceback.print_exc()
            errors.append(f"cemac_export_scratch: {e}")

    # 1g) Export CEMAC (with template)
    if not result_df.empty and not q_cemac.empty:
        print(f"\n  Export CEMAC (with template: {CEMAC_TEMPLATE})...")
        print(f"    Template existe : {CEMAC_TEMPLATE.exists()}")
        try:
            data_bytes = write_cemac_excel(
                result_df, q_cemac, POIDS_PIB,
                template_path=CEMAC_TEMPLATE,
            )
            print(f"    OK : {len(data_bytes)} bytes")

            # Vérifier le contenu
            wb = openpyxl.load_workbook(io.BytesIO(data_bytes))
            print(f"    Feuilles : {wb.sheetnames}")
            ws = wb["ICAE_Pays"]
            print(f"    ICAE_Pays max_row : {ws.max_row}")
            # Vérifier les formules
            r5_reb = ws.cell(5, 9).value
            r5_icae = ws.cell(5, 15).value
            print(f"    R5 col I (rebasé) : {str(r5_reb)[:60]}...")
            print(f"    R5 col O (ICAE CEMAC) : {str(r5_icae)[:60]}...")
            wb.close()
        except Exception as e:
            print(f"    ERREUR export template : {e}")
            traceback.print_exc()
            errors.append(f"cemac_export_template: {e}")

# =====================================================================
# PARTIE 2 : Test Module 2 recalc export end-to-end
# =====================================================================
print("\n\n" + "=" * 70)
print("PARTIE 2 : Test Module 2 recalc export end-to-end")
print("=" * 70)

# Utiliser le fichier RCA depuis Livrable_Final
test_code = "RCA"
fpath_recalc = os.path.join(LIVRABLE_DIR, f"ICAE_{test_code}_Consolide_now.xlsx")
if not os.path.exists(fpath_recalc):
    # Essayer le dossier Livrables
    fpath_recalc = str(CONSOLIDES / f"ICAE_{test_code}_Consolide.xlsx")

print(f"Fichier source : {fpath_recalc}")
print(f"Existe : {os.path.exists(fpath_recalc)}")

try:
    # 2a) Charger les données (comme Module 2)
    sheets = list_sheets(fpath_recalc)
    consignes = read_consignes(fpath_recalc) if "Consignes" in sheets else {"base_year": 2023}
    codif = read_codification(fpath_recalc) if "Codification" in sheets else pd.DataFrame()
    donnees = read_donnees_calcul(fpath_recalc)
    if not codif.empty:
        donnees = rename_columns_to_codes(donnees, codif)

    print(f"  Données chargées : {len(donnees)} lignes, {len(donnees.columns)-1} variables")
    print(f"  Dernière date : {pd.to_datetime(donnees['Date']).iloc[-1]}")

    # 2b) Simuler des prévisions (3 mois)
    horizon = 3
    last_date = pd.to_datetime(donnees["Date"]).max()
    fcst_dates = pd.date_range(last_date + pd.DateOffset(months=1), periods=horizon, freq="MS")
    print(f"  Dates prévision : {list(fcst_dates.strftime('%Y-%m'))}")

    new_rows = pd.DataFrame({"Date": fcst_dates})
    data_cols = [c for c in donnees.columns if c != "Date"]
    for col_name in data_cols:
        # Prévision simple : dernière valeur ± bruit
        last_val = donnees[col_name].dropna().iloc[-1] if donnees[col_name].notna().any() else 100
        new_rows[col_name] = [last_val * (1 + np.random.randn() * 0.02) for _ in range(horizon)]

    extended = pd.concat([donnees, new_rows], ignore_index=True)
    print(f"  Extended : {len(extended)} lignes ({len(donnees)} originales + {horizon} prévisions)")

    # 2c) Recalcul ICAE
    base_year = consignes.get("base_year", 2023)
    dates_c = pd.to_datetime(extended["Date"])
    base_mask_c = dates_c.dt.year == base_year
    base_idx = extended.index[base_mask_c]
    base_rows = range(base_idx[0], base_idx[-1] + 1) if len(base_idx) > 0 else range(108, 120)

    if not codif.empty and "Code" in codif.columns and "PRIOR" in codif.columns:
        priors = pd.Series(codif["PRIOR"].values, index=codif["Code"].values, dtype=float).fillna(0)
    else:
        priors = pd.Series(1.0, index=data_cols)

    results_icae = run_icae_pipeline(
        donnees=extended, priors=priors,
        base_year=base_year, base_rows=base_rows,
    )
    print(f"  ICAE recalculé : {len(results_icae['icae'])} obs")

    q = quarterly_mean(results_icae["icae"], results_icae["dates"])
    q["GA_Trim"] = calc_ga_trim(q["icae_trim"])
    q["GT_Trim"] = calc_gt_trim(q["icae_trim"])
    print(f"  Trimestriel : {len(q)} trimestres")

    # Contributions
    contrib_trim = None
    if "m" in results_icae["weights"] and not codif.empty:
        codif_for_contrib = codif
        contrib_trim = contributions_sectorielles_trim(
            results_icae["weights"]["m"], codif_for_contrib,
            results_icae["dates"],
        )
        if contrib_trim is not None:
            contrib_trim = normalize_contrib_to_ga(contrib_trim, q["GA_Trim"])
            print(f"  Contributions : {contrib_trim.shape}")

    # 2d) Export recalculé AVEC template
    print(f"\n  Export recalculé avec template...")
    wb_bytes = write_icae_recalc_output(
        donnees=extended,
        results=results_icae,
        quarterly=q,
        contrib_trim=contrib_trim,
        codification=codif,
        country_code=test_code,
        source_path=fpath_recalc,
    )
    print(f"    Taille : {len(wb_bytes)} bytes")

    # 2e) Vérifier le contenu du classeur exporté
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    print(f"    Feuilles : {wb.sheetnames}")

    ws_d = wb["Donnees_calcul"]
    print(f"    Donnees_calcul max_row : {ws_d.max_row}")
    print(f"    Attendu : {len(extended) + 1} (header + {len(extended)} données)")

    # Vérifier que les dernières lignes ont des valeurs
    for r in range(ws_d.max_row - 2, ws_d.max_row + 1):
        date_val = ws_d.cell(r, 1).value
        data_val = ws_d.cell(r, 2).value
        fill = ws_d.cell(r, 2).fill
        fill_color = fill.start_color.rgb if fill and fill.patternType else "aucun"
        print(f"    R{r}: date={date_val}, col2={data_val}, fill={fill_color}")

    if ws_d.max_row < len(extended) + 1:
        print(f"    ⚠️ PROBLÈME : max_row={ws_d.max_row} < attendu={len(extended)+1}")
        print(f"    Les séries ne sont PAS prolongées dans Donnees_calcul !")
        errors.append("Donnees_calcul non prolongé")
    else:
        print(f"    ✅ Donnees_calcul prolongé correctement")

    # Vérifier CALCUL_ICAE
    ws_c = wb["CALCUL_ICAE"]
    print(f"\n    CALCUL_ICAE max_row : {ws_c.max_row}")
    for r in range(ws_c.max_row - 2, ws_c.max_row + 1):
        date_val = ws_c.cell(r, 1).value
        tcs_val = ws_c.cell(r, 2).value
        is_formula = str(tcs_val)[:1] == "=" if tcs_val else False
        print(f"    R{r}: col_A={date_val}, col_B={str(tcs_val)[:60]}, formula={is_formula}")

    # Vérifier Resultats_Trim
    ws_rt = wb["Resultats_Trim"]
    print(f"\n    Resultats_Trim max_row : {ws_rt.max_row}")

    wb.close()

    # 2f) Export recalculé SANS template (fallback)
    print(f"\n  Export recalculé SANS template (fallback)...")
    wb_bytes2 = write_icae_recalc_output(
        donnees=extended,
        results=results_icae,
        quarterly=q,
        contrib_trim=contrib_trim,
        codification=codif,
        country_code=test_code,
        source_path=None,
    )
    print(f"    Taille : {len(wb_bytes2)} bytes")
    wb2 = openpyxl.load_workbook(io.BytesIO(wb_bytes2))
    print(f"    Feuilles : {wb2.sheetnames}")
    ws_d2 = wb2["Donnees_calcul"]
    print(f"    Donnees_calcul max_row : {ws_d2.max_row}")
    wb2.close()

except Exception as e:
    print(f"  ERREUR FATALE : {e}")
    traceback.print_exc()
    errors.append(f"Module 2 recalc: {e}")

# =====================================================================
# RÉSUMÉ
# =====================================================================
print("\n\n" + "=" * 70)
if errors:
    print(f"ÉCHECS : {len(errors)} erreur(s)")
    for e in errors:
        print(f"  - {e}")
else:
    print("TOUS LES TESTS PASSENT !")
print("=" * 70)
