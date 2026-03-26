"""Lecture des fichiers Excel consolidés ICAE."""
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path


def list_sheets(filepath) -> list:
    """Liste les feuilles d'un fichier Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def read_consignes(filepath, sheet="Consignes") -> dict:
    """Lit les métadonnées depuis la feuille Consignes."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]

    info = {}
    # Parcourir les lignes pour trouver les infos clés
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=2, values_only=True), start=1):
        r = row_idx
        label = row[0] if row else None
        val = row[1] if row and len(row) > 1 else None
        if r == 19 or (label and "ann" in str(label).lower() and "base" in str(label).lower()):
            try:
                info["base_year"] = int(val)
            except (TypeError, ValueError):
                pass
        if r == 20 or (label and "ligne" in str(label).lower() and "base" in str(label).lower()):
            if val and ":" in str(val):
                parts = str(val).split(":")
                info["base_rows_start"] = int(parts[0])
                info["base_rows_end"] = int(parts[1])

    wb.close()

    # Valeurs par défaut si non trouvées
    info.setdefault("base_year", 2023)
    info.setdefault("base_rows_start", 124)
    info.setdefault("base_rows_end", 135)

    return info


def read_codification(filepath, sheet="Codification") -> pd.DataFrame:
    """Lit la feuille Codification."""
    df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
    # Normaliser les colonnes
    col_map = {}
    for c in df.columns:
        cl = str(c).lower().strip()
        if "code" in cl:
            col_map[c] = "Code"
        elif "label" in cl or "libell" in cl:
            col_map[c] = "Label"
        elif "unit" in cl or "source" in cl:
            col_map[c] = "Unite_Source"
        elif "sect" in cl:
            col_map[c] = "Secteur"
        elif "prior" in cl:
            col_map[c] = "PRIOR"
        elif "stat" in cl:
            col_map[c] = "Statut"
    df = df.rename(columns=col_map)
    return df


def label_to_code_map(codification: pd.DataFrame) -> dict:
    """Crée un mapping Label→Code depuis la codification."""
    if "Code" in codification.columns and "Label" in codification.columns:
        return dict(zip(codification["Label"], codification["Code"]))
    return {}


def rename_columns_to_codes(donnees: pd.DataFrame,
                            codification: pd.DataFrame) -> pd.DataFrame:
    """Renomme les colonnes de Donnees_calcul (Labels) en Codes."""
    l2c = label_to_code_map(codification)
    if not l2c:
        return donnees
    rename_map = {}
    for col in donnees.columns:
        if col != "Date" and col in l2c:
            rename_map[col] = l2c[col]
    return donnees.rename(columns=rename_map)


def read_donnees_calcul(filepath, sheet="Donnees_calcul") -> pd.DataFrame:
    """Lit les données de calcul (séries temporelles brutes)."""
    df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
    # La première colonne est la date
    date_col = df.columns[0]
    df = df.rename(columns={date_col: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    # Supprimer les lignes sans date valide (lignes parasites en fin de feuille)
    df = df.dropna(subset=["Date"]).reset_index(drop=True)
    # Supprimer les colonnes sans en-tête (Unnamed)
    df = df[[c for c in df.columns if not str(c).startswith("Unnamed")]]
    return df


def read_calcul_icae(filepath, sheet="CALCUL_ICAE") -> dict:
    """Lit la feuille CALCUL_ICAE pour extraire les paramètres existants."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]

    # Lire les 15 premières lignes (en-têtes et paramètres)
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        rows.append(list(row))

    # Ligne 1 : Inv_Ecart_Type
    inv_et = rows[0] if rows else []
    # Ligne 2 : Pondération (a)
    pond_a = rows[1] if len(rows) > 1 else []
    # Ligne 9 : PRIOR MEDIAN
    prior_median = rows[8] if len(rows) > 8 else []
    # Ligne 12 : Pondération finale
    pond_finale = rows[11] if len(rows) > 11 else []
    # Ligne 14 : Codes variables
    codes = rows[13] if len(rows) > 13 else []

    # Lire les données complètes
    data_rows = []
    for row in ws.iter_rows(min_row=16, values_only=True):
        data_rows.append(list(row))

    wb.close()

    return {
        "inv_et": inv_et,
        "pond_a": pond_a,
        "prior_median": prior_median,
        "pond_finale": pond_finale,
        "codes": codes,
        "data_rows": data_rows,
    }


def read_resultats_trim(filepath, sheet="Resultats_Trim") -> pd.DataFrame:
    """Lit les résultats trimestriels."""
    df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
    return df


def read_contrib(filepath, sheet="Contrib") -> pd.DataFrame:
    """Lit les contributions."""
    df = pd.read_excel(filepath, sheet_name=sheet, engine="openpyxl")
    return df


# ────────────────────────────────────────────────────────────────────────────
# Fichier CEMAC
# ────────────────────────────────────────────────────────────────────────────
def read_cemac_poids(filepath, sheet="Poids_PIB") -> pd.DataFrame:
    """Lit les poids PIB depuis le fichier CEMAC."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        rows.append(list(row))
    wb.close()

    if not rows:
        return pd.DataFrame()

    # Première ligne = en-têtes
    headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=headers)
    return df


def read_cemac_icae_pays(filepath, sheet="ICAE_Pays") -> pd.DataFrame:
    """Lit l'ICAE par pays depuis le fichier CEMAC."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        rows.append(list(row))
    wb.close()

    if not rows:
        return pd.DataFrame()

    headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=headers)
    return df


def read_cemac_icae_trim(filepath, sheet="ICAE_Trimestriel") -> pd.DataFrame:
    """Lit l'ICAE trimestriel depuis le fichier CEMAC."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        rows.append(list(row))
    wb.close()

    if not rows:
        return pd.DataFrame()

    headers = [str(h) if h else f"Col{i}" for i, h in enumerate(rows[0])]
    df = pd.DataFrame(rows[1:], columns=headers)
    return df


# ────────────────────────────────────────────────────────────────────────────
# Lecture complète d'un fichier pays
# ────────────────────────────────────────────────────────────────────────────
def load_country_file(filepath) -> dict:
    """Charge un fichier ICAE pays complet."""
    filepath = Path(filepath)
    sheets = list_sheets(filepath)

    result = {"filepath": filepath, "sheets": sheets}

    if "Consignes" in sheets:
        result["consignes"] = read_consignes(filepath)
    if "Codification" in sheets:
        result["codification"] = read_codification(filepath)
    if "Donnees_calcul" in sheets:
        result["donnees_calcul"] = read_donnees_calcul(filepath)
    if "CALCUL_ICAE" in sheets:
        result["calcul_icae"] = read_calcul_icae(filepath)
    if "Resultats_Trim" in sheets:
        result["resultats_trim"] = read_resultats_trim(filepath)
    if "Contrib" in sheets:
        result["contrib"] = read_contrib(filepath)

    return result


def load_cemac_file(filepath) -> dict:
    """Charge le fichier ICAE CEMAC complet."""
    filepath = Path(filepath)
    sheets = list_sheets(filepath)

    result = {"filepath": filepath, "sheets": sheets}

    if "Poids_PIB" in sheets:
        result["poids_pib"] = read_cemac_poids(filepath)
    if "ICAE_Pays" in sheets:
        result["icae_pays"] = read_cemac_icae_pays(filepath)
    if "ICAE_Trimestriel" in sheets:
        result["icae_trim"] = read_cemac_icae_trim(filepath)

    return result
