"""Écriture de fichiers Excel avec formules et valeurs."""
import io
import copy
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# Styles
HIST_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
HIST_FONT = Font(color="1F4E79")
FCST_FILL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
FCST_FONT = Font(color="E26B0A", bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_icae_output(source_path, results: dict, country_code: str) -> bytes:
    """
    Crée un fichier _OUT.xlsx à partir du source avec résultats injectés.
    Resultats_Trim contient des formules Excel d'agrégation référençant CALCUL_ICAE.

    Returns : bytes du fichier Excel
    """
    wb = openpyxl.load_workbook(source_path)

    # ── Mettre à jour CALCUL_ICAE ─────────────────────────────────────────
    icae_col_letter = None
    data_start_row = 16  # les données commencent à la ligne 16 dans CALCUL_ICAE

    if "CALCUL_ICAE" in wb.sheetnames:
        ws = wb["CALCUL_ICAE"]
        icae = results.get("icae")

        if icae is not None:
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                h = ws.cell(row=15, column=col).value
                if h and "ICAE" in str(h):
                    icae_col_letter = get_column_letter(col)
                    for i, val in enumerate(icae.values):
                        if not np.isnan(val):
                            r = data_start_row + i
                            ws.cell(row=r, column=col, value=round(val, 6))

    # ── Mettre à jour Resultats_Trim avec formules Excel ──────────────────
    if "Resultats_Trim" in wb.sheetnames and "quarterly" in results:
        ws_rt = wb["Resultats_Trim"]
        q_data = results["quarterly"]
        dates = results.get("dates")

        # Construire la correspondance trimestre → lignes dans CALCUL_ICAE
        if dates is not None and icae_col_letter is not None:
            dates_ts = pd.to_datetime(dates)
            quarter_groups = {}
            for i, d in enumerate(dates_ts):
                qkey = f"{d.year}T{(d.month - 1) // 3 + 1}"
                if qkey not in quarter_groups:
                    quarter_groups[qkey] = []
                quarter_groups[qkey].append(data_start_row + i)

            # En-têtes pour Resultats_Trim
            headers = ["Trimestre", "ICAE_Trim", "GA_Trim", "GT_Trim"]
            for j, h in enumerate(headers):
                cell = ws_rt.cell(row=1, column=j + 1, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER

            # Écrire les formules pour chaque trimestre
            for idx, (_, row) in enumerate(q_data.iterrows()):
                r = idx + 2
                trim_label = row.get("trimestre", str(row.get("quarter", "")))
                ws_rt.cell(row=r, column=1, value=trim_label).border = THIN_BORDER

                rows_in_q = quarter_groups.get(trim_label, [])

                if rows_in_q and icae_col_letter:
                    # ICAE_Trim = AVERAGE(CALCUL_ICAE!col:col) sur les 3 mois
                    refs = ",".join(
                        f"CALCUL_ICAE!{icae_col_letter}{rr}" for rr in rows_in_q
                    )
                    formula_avg = f"=AVERAGE({refs})"
                    ws_rt.cell(row=r, column=2, value=formula_avg).border = THIN_BORDER

                    # GA_Trim = (ICAE_T / ICAE_{T-4} - 1) * 100
                    if idx >= 4:
                        r_prev_year = idx - 4 + 2
                        col_b = get_column_letter(2)
                        formula_ga = f"=({col_b}{r}/{col_b}{r_prev_year}-1)*100"
                        ws_rt.cell(row=r, column=3, value=formula_ga).border = THIN_BORDER
                    else:
                        ws_rt.cell(row=r, column=3).border = THIN_BORDER

                    # GT_Trim = (ICAE_T / ICAE_{T-1} - 1) * 100
                    if idx >= 1:
                        r_prev = idx - 1 + 2
                        formula_gt = f"=({col_b}{r}/{col_b}{r_prev}-1)*100"
                        ws_rt.cell(row=r, column=4, value=formula_gt).border = THIN_BORDER
                    else:
                        ws_rt.cell(row=r, column=4).border = THIN_BORDER
                else:
                    # Pas de correspondance, écrire les valeurs directement
                    for j, col_name in enumerate(["icae_trim", "GA_Trim", "GT_Trim"]):
                        val = row.get(col_name)
                        if val is not None and not (isinstance(val, float) and np.isnan(val)):
                            ws_rt.cell(row=r, column=j + 2, value=round(float(val), 4))
                        ws_rt.cell(row=r, column=j + 2).border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


def write_previsions_excel(donnees: pd.DataFrame, previsions: dict,
                           backtesting: dict, recommandations: dict,
                           scenarios: dict = None,
                           edited: pd.DataFrame = None) -> bytes:
    """Génère le fichier Excel de prévisions."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # Feuille Donnees
        donnees.to_excel(writer, sheet_name="Donnees", index=False)

        # Feuille Previsions
        prev_rows = []
        for var, methods in previsions.items():
            for method, values in methods.items():
                for i, v in enumerate(values):
                    prev_rows.append({
                        "Variable": var, "Methode": method,
                        "Mois": i + 1, "Valeur": v,
                    })
        if prev_rows:
            pd.DataFrame(prev_rows).to_excel(
                writer, sheet_name="Previsions", index=False)

        # Feuille Backtesting
        bt_rows = []
        for var, methods in backtesting.items():
            row = {"Variable": var}
            for method, metrics in methods.items():
                row[f"{method}_MAPE"] = metrics.get("mape", np.nan)
                row[f"{method}_RMSE"] = metrics.get("rmse", np.nan)
            bt_rows.append(row)
        if bt_rows:
            pd.DataFrame(bt_rows).to_excel(
                writer, sheet_name="Backtesting", index=False)

        # Feuille Recommandations
        rec_rows = []
        for var, info in recommandations.items():
            rec_rows.append({
                "Variable": var,
                "Methode_recommandee": info["method"],
                "MAPE": info.get("mape", np.nan),
            })
        if rec_rows:
            pd.DataFrame(rec_rows).to_excel(
                writer, sheet_name="Recommandations", index=False)

        # Feuille Scenarios
        if scenarios is not None:
            scenarios.to_excel(writer, sheet_name="ICAE_Scenarios", index=False)

        # Feuille Previsions_Editees
        if edited is not None:
            edited.to_excel(writer, sheet_name="Previsions_Editees",
                            index=False)

    buf.seek(0)
    return buf.getvalue()


def write_nowcast_excel(pib_q: pd.Series, results: dict,
                        params: dict = None) -> bytes:
    """Génère le fichier Excel de résultats Nowcast."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # PIB et Nowcasts
        df = pd.DataFrame({"PIB_observe": pib_q})
        for name, r in results.items():
            fc = r["forecast"]
            df[name] = fc.reindex(df.index)
        df.to_excel(writer, sheet_name="PIB_and_Nowcasts")

        # Performance
        perf_rows = []
        for name, r in results.items():
            m = r["metrics"]
            perf_rows.append({
                "Modele": name,
                "RMSE_in": m["in_sample"].get("rmse", np.nan),
                "MAE_in": m["in_sample"].get("mae", np.nan),
                "MAPE_in": m["in_sample"].get("mape", np.nan),
                "RMSE_out": m["out_sample"].get("rmse", np.nan),
                "MAE_out": m["out_sample"].get("mae", np.nan),
                "MAPE_out": m["out_sample"].get("mape", np.nan),
                "Correlation": r.get("correlation", np.nan),
            })
        pd.DataFrame(perf_rows).to_excel(
            writer, sheet_name="Performance", index=False)

        # GA
        if len(df) > 4:
            ga_df = pd.DataFrame(index=df.index)
            for col in df.columns:
                ga_df[f"GA_{col}"] = (df[col] / df[col].shift(4) - 1) * 100
            ga_df.to_excel(writer, sheet_name="Glissement_Annuel")

        # Paramètres
        if params:
            pd.DataFrame([params]).to_excel(
                writer, sheet_name="Parametres", index=False)

    buf.seek(0)
    return buf.getvalue()


def write_cemac_excel(df_monthly: pd.DataFrame,
                      df_quarterly: pd.DataFrame,
                      poids: dict) -> bytes:
    """Génère le fichier Excel CEMAC."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # Poids PIB
        poids_df = pd.DataFrame([
            {"Code": k, "Poids": v} for k, v in poids.items()
        ])
        poids_df.to_excel(writer, sheet_name="Poids_PIB", index=False)

        # ICAE Pays (mensuel)
        df_monthly.to_excel(writer, sheet_name="ICAE_Pays", index=True)

        # ICAE Trimestriel
        df_quarterly.to_excel(writer, sheet_name="ICAE_Trimestriel",
                              index=False)

    buf.seek(0)
    return buf.getvalue()
